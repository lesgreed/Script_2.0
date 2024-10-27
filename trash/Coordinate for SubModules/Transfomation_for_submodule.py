import openpyxl
import os
import numpy as np 
from scipy.spatial.transform import Rotation as R


def delete(file_to_delete):

 if os.path.exists(file_to_delete):
    os.remove(file_to_delete)
    print(f"File {file_to_delete} deleted.")
 else:
    print(f"File {file_to_delete} not found.")  
    
def create_new_file(input_file_path, output_file_path):
    

 input_workbook = openpyxl.load_workbook(input_file_path)
 input_sheet = input_workbook.active

 output_workbook = openpyxl.Workbook()
 output_sheet = output_workbook.active

 for row in input_sheet.iter_rows():
    output_sheet.append([cell.value for cell in row])
 return output_sheet, output_workbook

def Module(output_sheet):
 for row in output_sheet.iter_rows(min_row=4, max_row=output_sheet.max_row, min_col=3, max_col=3):
    cell_value = row[0].value
    row_number = row[0].row
    
    
    new_values = new_coordinates_for_Moduls(row_number, output_sheet, cell_value)
    for i in range(len(new_values)):
        output_sheet.cell(row=row_number, column=(i+5)).value = new_values[i]
           
 return output_file_path
 

def translate_coord(X, Y, Z):

    R = np.sqrt(X**2 + Y**2)
    if X==0:
        Phi_0 = np.pi/2
    else:   
        Phi_0 = np.arctan(Y/X)
    
    Phi = np.degrees(Phi_0)
    Z_r = Z
    return R, Phi, Z_r

def back_translate_coord(R, Phi, Z):
    angle = np.radians(Phi)
    x = R * np.cos(angle)
    y = R * np.sin(angle)
    z = Z
    return x, y, z 
    
 
def new_coordinates_for_Moduls(row_number, output_sheet, cell_value):
    X_1 = output_sheet[row_number][4].value
    Y_1 = output_sheet[row_number][5].value
    Z_1 = output_sheet[row_number][6].value
    
    X_2 = output_sheet[row_number][7].value
    Y_2 = output_sheet[row_number][8].value
    Z_2 = output_sheet[row_number][9].value
    
    X_3 = output_sheet[row_number][10].value
    Y_3 = output_sheet[row_number][11].value
    Z_3 = output_sheet[row_number][12].value
    
    X_4 = output_sheet[row_number][13].value
    Y_4 = output_sheet[row_number][14].value
    Z_4 = output_sheet[row_number][15].value
    
    R_1, Phi_1, Z_1 =  translate_coord(X_1, Y_1, Z_1)
    R_2, Phi_2, Z_2 =  translate_coord(X_2, Y_2, Z_2)
    R_3, Phi_3, Z_3 =  translate_coord(X_3, Y_3, Z_3)
    R_4, Phi_4, Z_4 =  translate_coord(X_4, Y_4, Z_4)
    
    Phi_1 =   Phi_1  + 72* (cell_value -1) 
    Phi_2 =   Phi_2  + 72* (cell_value -1) 
    Phi_3 =   Phi_3  + 72* (cell_value -1) 
    Phi_4 =   Phi_4  + 72* (cell_value -1) 
    
    
    X_1, Y_1, Z_1 = back_translate_coord(R_1, Phi_1, Z_1)
    X_2, Y_2, Z_2 = back_translate_coord(R_2, Phi_2, Z_2)
    X_3, Y_3, Z_3 = back_translate_coord(R_3, Phi_3, Z_3)
    X_4, Y_4, Z_4 = back_translate_coord(R_4, Phi_4, Z_4)
    
    output_sheet[row_number][4].value = X_1
    output_sheet[row_number][5].value = Y_1
    output_sheet[row_number][6].value = Z_1
            
    output_sheet[row_number][7].value = X_2
    output_sheet[row_number][8].value = Y_2
    output_sheet[row_number][9].value = Z_2
            
    output_sheet[row_number][10].value = X_3
    output_sheet[row_number][11].value = Y_3
    output_sheet[row_number][12].value = Z_3
            
    output_sheet[row_number][13].value = X_4
    output_sheet[row_number][14].value = Y_4
    output_sheet[row_number][15].value = Z_4

    return [output_sheet[row_number][4].value, output_sheet[row_number][5].value, output_sheet[row_number][6].value,
            output_sheet[row_number][7].value, output_sheet[row_number][8].value, output_sheet[row_number][9].value,
            output_sheet[row_number][10].value, output_sheet[row_number][11].value, output_sheet[row_number][12].value,
            output_sheet[row_number][13].value, output_sheet[row_number][14].value, output_sheet[row_number][15].value]
    
def rotate_point_around_arbitrary_axis(point, axis_point1, axis_point2):
    angle_deg = 180
    angle_rad = np.radians(angle_deg)

    point = np.array(point)
    axis = np.array(axis_point2) - np.array(axis_point1)
    

    axis_length = np.linalg.norm(axis)
    axis_normalized = axis / axis_length
    

    rotation = R.from_rotvec(axis_normalized * angle_rad)
    

    rotated_point = rotation.apply(point - np.array(axis_point1)) + np.array(axis_point1)
    
    return np.array(rotated_point)


def submodule(output_sheet):
    for row in output_sheet.iter_rows(min_row=4, max_row=output_sheet.max_row, min_col=3, max_col=3):
       module_value = row[0].value
       row_number = row[0].row
       submodule_value = output_sheet[row_number][3].value

       if submodule_value == "rotation":
        submodule_value = 1
       print(submodule_value)
       if submodule_value == 0:
           new_values2 = new_coordinates_for_SUBModuls(row_number, output_sheet, module_value, submodule_value)
           for i in range(len(new_values2)):
               output_sheet.cell(row=row_number, column=(i+5)).value = new_values2[i]
    return output_file_path
    

def new_coordinates_for_SUBModuls(row_number, output_sheet, module_value, submodule_value):
    X_1 = output_sheet[row_number][4].value
    Y_1 = output_sheet[row_number][5].value
    Z_1 = output_sheet[row_number][6].value
    
    X_2 = output_sheet[row_number][7].value
    Y_2 = output_sheet[row_number][8].value
    Z_2 = output_sheet[row_number][9].value
    
    X_3 = output_sheet[row_number][10].value
    Y_3 = output_sheet[row_number][11].value
    Z_3 = output_sheet[row_number][12].value
    
    X_4 = output_sheet[row_number][13].value
    Y_4 = output_sheet[row_number][14].value
    Z_4 = output_sheet[row_number][15].value
    
    
    X_Zero = 0
    Y_Zero = 0 
    Z_Zero = 0
    
    
    R_two = 1
    Phi_two = 36 + 72 * (module_value-1)
    Z_two = 0
    
    X_two, Y_two, Z_two =  back_translate_coord(R_two,  Phi_two, Z_two)
    
    

    axis_point1 = np.array([X_Zero, Y_Zero, Z_Zero])
    axis_point2 = np.array([X_two, Y_two, Z_two])
    
    
    point1 = np.array([X_1, Y_1, Z_1])
    point2 = np.array([X_2, Y_2, Z_2])
    point3 = np.array([X_3, Y_3, Z_3])
    point4 = np.array([X_4, Y_4, Z_4])
    
    rotated_point1 = rotate_point_around_arbitrary_axis(point1, axis_point1, axis_point2)
    rotated_point2 = rotate_point_around_arbitrary_axis(point2, axis_point1, axis_point2)
    rotated_point3 = rotate_point_around_arbitrary_axis(point3, axis_point1, axis_point2)
    rotated_point4 = rotate_point_around_arbitrary_axis(point4, axis_point1, axis_point2)
    
    output_sheet[row_number][4].value = rotated_point1[0]
    output_sheet[row_number][5].value = rotated_point1[1]
    output_sheet[row_number][6].value = rotated_point1[2]
            
    output_sheet[row_number][7].value = rotated_point2[0]
    output_sheet[row_number][8].value = rotated_point2[1]
    output_sheet[row_number][9].value = rotated_point2[2]
            
    output_sheet[row_number][10].value = rotated_point3[0]
    output_sheet[row_number][11].value = rotated_point3[1]
    output_sheet[row_number][12].value = rotated_point3[2]
            
    output_sheet[row_number][13].value = rotated_point4[0]
    output_sheet[row_number][14].value = rotated_point4[1]
    output_sheet[row_number][15].value = rotated_point4[2]

    return [output_sheet[row_number][4].value, output_sheet[row_number][5].value, output_sheet[row_number][6].value,
            output_sheet[row_number][7].value, output_sheet[row_number][8].value, output_sheet[row_number][9].value,
            output_sheet[row_number][10].value, output_sheet[row_number][11].value, output_sheet[row_number][12].value,
            output_sheet[row_number][13].value, output_sheet[row_number][14].value, output_sheet[row_number][15].value]
    







if __name__ == "__main__":
 os.chdir('Input_data\PreInput\Coordinates for modules')
 print("Current working directory:", os.getcwd())
        
 file_to_delete = 'Result_Tranformation_for_submodule.xlsx'
 input_file_path = 'Result_Tranformation_for_module.xlsx'
 output_file_path = 'Result_Tranformation_for_submodule2.xlsx'
 


 #delete(file_to_delete)    
  
 
 output_sheet, output_workbook = create_new_file(input_file_path, output_file_path)
 
 
 output_file_path = submodule(output_sheet)
 
 output_workbook.save(output_file_path)
 output_workbook.close()
 



 
