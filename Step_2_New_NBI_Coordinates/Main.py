import Tools.Function_for_DATA_angel as FuD 
import coordinate_output_NBI_and_ports as Cout
import Plot_3d as pl3d
import Ports_and_point_on_line_NBI as Port
from openpyxl import Workbook
import os
import numpy as np




def write_data_to_excel(output_sheet, headers, data):
    # Write column headers
    for i, header in enumerate(headers, start=5):
        output_sheet.cell(row=5, column=i).value = header

    # Write data in separate columns
    for col_idx, column_data in enumerate(data, start=5):
        for row_idx, value in enumerate(column_data, start=6):
            output_sheet.cell(row=row_idx, column=col_idx).value = value

    return output_sheet





if __name__ == "__main__":

#Create Full DATA poinst in XYZ coordinate
    Phi, R_phi, Z_phi = FuD.read_data('Tools/data.txt')
    R_x_all, R_y_all, Z_all = FuD.all_point(Phi)
    
    
    
#Downnload Input data of coordinates ports and NBI
    NBI_X, NBI_Y, NBI_Z, NBI_uvec_X, NBI_uvec_Y, NBI_uvec_Z = Cout.NBI()
    P_1, P_2, P_3, P_unit_vector = Cout.Ports()   
    
    
#Find new NBI coordinates    
    NBI_new_start = Port.find_new_NBI_start() 
    NBI_new_end = Port.find_new_NBI_end(NBI_new_start)



        
# Define your data
    headers = ["NBI_new_start[x]", "NBI_new_start[y]", "NBI_new_start[z]", "NBI_new_end[x]", "NBI_new_end[y]", "NBI_new_end[z]"]
    data = [NBI_new_start[0], NBI_new_start[1], NBI_new_start[2], NBI_new_end[0], NBI_new_end[1], NBI_new_end[2]]

# Create a new Excel workbook and sheet
    workbook = Workbook()
    output_sheet = workbook.active

# Write each data array to a separate column
    output_sheet = write_data_to_excel(output_sheet, headers, data)

# Save the workbook
    current_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(current_dir, 'Result_NBI')
    output_file_path = os.path.join(output_dir, 'New_coordinate_NBI_GT_21.04.xlsx') 
    workbook.save(output_file_path)



#Check with graph 
    pl3d.plot_3d(R_x_all, R_y_all, Z_all, NBI_new_start, NBI_new_end)    