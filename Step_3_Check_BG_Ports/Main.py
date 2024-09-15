import Tools.Function_for_DATA_angel as FuD 
import coordinate_output_NBI_and_ports as Cout
import Plot_3d as pl3d
import Check_BG as Ch
from openpyxl import Workbook
import os



def write_data_to_excel(output_sheet, headers, data):
    # Write column headers
    for i, header in enumerate(headers, start=5):
        output_sheet.cell(row=5, column=i).value = header

    # Write data in separate columns
    for col_idx, column_data in enumerate(data, start=5):
        for row_idx, value in enumerate(column_data, start=6):
            output_sheet.cell(row=row_idx, column=col_idx).value = value

    return output_sheet





if __name__ == "__main__":

#Create Full DATA poinst in XYZ coordinate
    Phi, R_phi, Z_phi = FuD.read_data('Tools/data.txt')
    R_x_all, R_y_all, Z_all = FuD.all_point(Phi)
    
#Downnload Input data of coordinates ports and NBI
    P_1_new, P_2_new = Cout.new_Ports()
    NBI_start, NBI_end =Cout.new_NBI()
    
#Check BG ports     
    Good_Ports = Ch.check()
    
    
    
# Define your data
    headers = ["Good_Ports_Nubmer_Ports", "Good_Ports_Nubmer_NBI"]
    data = [Good_Ports[0][0], Good_Ports[0][1]]

# Create a new Excel workbook and sheet
    workbook = Workbook()
    output_sheet = workbook.active

# Write each data array to a separate column
    output_sheet = write_data_to_excel(output_sheet, headers, data)

# Save the workbook
    current_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(current_dir, 'Result_BG_Ports')
    output_file_path = os.path.join(output_dir, 'Number_Good_Ports_and_NBI_14_03.xlsx') 
    workbook.save(output_file_path)



#Check with graph 
    pl3d.plot_3d(R_x_all, R_y_all, Z_all, P_1_new, P_2_new, Good_Ports, NBI_start, NBI_end)    